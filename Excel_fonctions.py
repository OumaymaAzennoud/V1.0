from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
import os


# ----------------------------------------------------------------------------------------------------------------------------------------------
# Global Variables
# ----------------------------------------------------------------------------------------------------------------------------------------------
#CSV2TABLE-------------------
Col_time=1
Col_voltage=2
Col_current = 3
Col_power = 4
#Charging IOP----------------
Col_Nlog=3
Col_Type=4
Col_ChargingType=5
Col_USBVersion=6
Col_EDTRef=7
Col_TypeOfProduct=8
Col_Brand=9
Col_Model=10
Col_MaxCurrent=11
Col_Led=12
Col_PrechargeCrurrent=13
Col_PrechargeEndVoltage=14
Col_PrechargeDuration=15
Col_NormalChargeCurrent=16
Col_NormalChargePower=17
Col_PrechargeExitDuration=18
Col_CurretAfterFlip=19
Col_PowerAfterFlip=20
Col_VoltageNormalCharge=21
Col_TotalChargeDuration=22
Col_PassRemarkFail=23
Col_Sample=24
Col_Cable=25
Col_Comment=26
#-------------------------------
Log_line_number=16
IOP_file="Template"
Table_file="CSV2Table"
precharge_start=0.05
precharge_end=0.2
# ----------------------------------------------------------------------------------------------------------------------------------------------
# Workspaces and worksheets
# ----------------------------------------------------------------------------------------------------------------------------------------------
wb1 = load_workbook(
    'C:/Users/oumayma.azennoud/OneDrive - Eurofins Digital Testing International/Desktop/Python/Template.xlsx')
ws1 = wb1['Charging IOP']

wb2 = load_workbook(
    'C:/Users/oumayma.azennoud/OneDrive - Eurofins Digital Testing International/Desktop/Python/CSV2Table.xlsx')
ws2 = wb2['Sheet1']
# ----------------------------------------------------------------------------------------------------------------------------------------------


#----------------------------------------------------------------------------------------------------------------------------------------------
#Import CSV file and copy it into CSV2Table Excel file.
#Arguments: Log_to_import : This is the number of the log & the name of the CSV file, Please ensure that all the Logs names are numbers
#----------------------------------------------------------------------------------------------------------------------------------------------
def Import_CSV(Log_to_import):
    path = os.path.dirname(r'C:/Users/oumayma.azennoud/OneDrive - Eurofins Digital Testing International/Desktop/Excel (1)\Excel/Excel/Logs/')
    csv_file = os.path.join(path, str(Log_to_import) + '.csv')
    read_file = pd.read_csv(csv_file)
    read_file.to_excel(
        r'C:/Users/oumayma.azennoud/OneDrive - Eurofins Digital Testing International/Desktop/Python/CSV2Table.xlsx',
        index=None, header=True)
#----------------------------------------------------------------------------------------------------------------------------------------------
#Fill a cell with data, data can be everything
#Arguments: Row : Line of the cell
#Arguments: Col : Column of the cell
#Arguments: Value : The data you want to put in the Cell
#Arguments: Sheet : the "Sheet" name
#Arguments: File : the "File" name
#----------------------------------------------------------------------------------------------------------------------------------------------

def FillCell(Row,Col,Value,Sheet,File):
    path=os.path.dirname(r'C:/Users/oumayma.azennoud/OneDrive - Eurofins Digital Testing International/Desktop/Python/')
    Xl_File = os.path.join(path, str(File) + '.xlsx')
    wb = load_workbook(Xl_File)
    ws = wb[str(Sheet)]
    char = get_column_letter(Col)
    ws[char + str(Row)].value = Value
    wb.save(Xl_File)

#----------------------------------------------------------------------------------------------------------------------------------------------
#Get date from a cell
#Arguments: Row : Line of the cell
#Arguments: Col : Column of the cell
#Arguments: Sheet : the "Sheet" name
#Arguments: File : the "File" name
#----------------------------------------------------------------------------------------------------------------------------------------------
def GetCellValue(Row, Col,Sheet,File):
    path=os.path.dirname(r'C:/Users/oumayma.azennoud/OneDrive - Eurofins Digital Testing International/Desktop/Python/')
    Xl_File = os.path.join(path, str(File) + '.xlsx')
    wb = load_workbook(Xl_File)
    ws = wb[str(Sheet)]
    char = get_column_letter(Col)
    return ws[char+str(Row)].value
#----------------------------------------------------------------------------------------------------------------------------------------------
#Calculate the number of cells containing data in a Row/Column
#Arguments: df : Keep it df, nothing to enter
#Arguments: Sheet name : "Sheet name"
#----------------------------------------------------------------------------------------------------------------------------------------------
def get_max_row_column(df, sheet_name):
    global max_row
    global max_col
    max_row = 1
    max_col = 1
    for sh_name, sh_content in df.items():
        if sh_name == sheet_name:
            max_row = len(sh_content) + 1
            max_col = len(sh_content.columns)
            break
    coordinates = {'max_row': max_row, 'max_col': max_col}
    return coordinates
df = pd.read_excel('C:/Users/oumayma.azennoud/OneDrive - Eurofins Digital Testing International/Desktop/Python/CSV2Table.xlsx', sheet_name=None)
max_row = get_max_row_column(df, 'Sheet1')['max_row']
max_col = get_max_row_column(df, 'Sheet1')['max_col']

#----------------------------------------------------------------------------------------------------------------------------------------------
#Get the colum
#Arguments: char : enter the column number or the column indicator
#----------------------------------------------------------------------------------------------------------------------------------------------
def GetCol(Col) :
    char = get_column_letter(Col)
    return char



#----------------------------------------------------------------------------------------------------------------------------------------------
#Fill in a line in IOP Sheet
#Arguments: Log_line_number : enter the line or the log you want to fill in
#----------------------------------------------------------------------------------------------------------------------------------------------
def Fill_IOP_row(Log_line_number):
    Max_current_ph1 = None
    duration_ph1 = None
    for i in range(2, max_row + 1, 1):
        if abs(ws2[GetCol(Col_current) + str(i)].value) > precharge_start:
            EndPh1 = i
            # duration calculation
            duration_ph1 = abs(ws2[GetCol(Col_time) + str(i)].value)
            duration_ph1 /= 1000000
            break
        # Current calculation
        current_ph1 = 0 if abs(float(ws2[GetCol(Col_current) + str(i)].value)) == None else abs(
            float(ws2[GetCol(Col_current) + str(i)].value))
        current_ph1 *= 1000
        if Max_current_ph1 is None or current_ph1 > Max_current_ph1: Max_current_ph1 = current_ph1

    Max_current_ph2 = None
    Max_power_ph2 = None  # initialize the variable
    for i in range(EndPh1, max_row + 1,
                   1):  # From the start of the phase (end of last phase) to the last cell with data in the column, the loop will break before that as soon as the next phase is detected
        if abs(ws2[GetCol(Col_current) + str(i)].value) > precharge_end:  # detect the next phase
            EndPh2 = i  # memorise the start row of the next phase
            # duration calculation
            duration_ph2 = abs(ws2[GetCol(Col_time) + str(i)].value) / 1000000
            duration_ph2 = duration_ph2 - duration_ph1
            break
        # Current calculation
        current_ph2 = 0 if abs(float(ws2[GetCol(Col_current) + str(i)].value)) == None else abs(
            float(ws2[GetCol(Col_current) + str(i)].value))  # to prevent the program to tak a "None" value
        current_ph2 *= 1000  # convert into mA

        if Max_current_ph2 is None or current_ph2 > Max_current_ph2: Max_current_ph2 = current_ph2  # select the max current value of the phase
    FillCell(12 + Log_line_number, Col_PrechargeCrurrent, Max_current_ph2, "Charging IOP",
                IOP_file)  # Fill the curret cell in Charging IOP sheet with Max_ph2
    FillCell(12 + Log_line_number, Col_PrechargeDuration, int(duration_ph2), "Charging IOP", IOP_file)

    Max_current_ph3 = None
    Max_power_ph3 = None
    Max_voltage_ph3 = None
    for i in range(EndPh2, max_row + 1, 1):
        if abs(ws2[GetCol(Col_current) + str(i)].value) < precharge_start:
            EndPh3 = i
            # duration calculation
            duration_ph3 = abs(ws2[GetCol(Col_time) + str(i)].value) / 1000000
            duration_ph3 = duration_ph3 - duration_ph2
            break
        # Current calculation
        current_ph3 = 0 if abs(float(ws2[GetCol(Col_current) + str(i)].value)) == None else abs(
            float(ws2[GetCol(Col_current) + str(i)].value))
        current_ph3 *= 1000
        if Max_current_ph3 is None or current_ph3 > Max_current_ph3: Max_current_ph3 = current_ph3
        # Power calculation
        power_ph3 = 0 if abs(float(ws2[GetCol(Col_power) + str(i)].value)) == None else abs(
            float(ws2[GetCol(Col_power) + str(i)].value))  # to prevent the program to tak a "None" value
        if Max_power_ph3 is None or power_ph3 > Max_power_ph3: Max_power_ph3 = power_ph3
        # Voltage calculation
        voltage_ph3 = 0 if abs(float(ws2[GetCol(Col_voltage) + str(i)].value)) == None else abs(
            float(ws2[GetCol(Col_voltage) + str(i)].value))  # to prevent the program to tak a "None" value
        if Max_voltage_ph3 is None or voltage_ph3 > Max_voltage_ph3: Max_voltage_ph3 = voltage_ph3
    FillCell(12 + Log_line_number, Col_NormalChargeCurrent, Max_current_ph3, "Charging IOP", IOP_file)
    FillCell(12 + Log_line_number, Col_NormalChargePower, Max_power_ph3, "Charging IOP",
                IOP_file)  # Fill the power cell in Charging IOP sheet with Max_ph2
    FillCell(12 + Log_line_number, Col_VoltageNormalCharge, Max_voltage_ph3, "Charging IOP",
                IOP_file)  # Fill the voltage cell in Charging IOP sheet with Max_ph2
    FillCell(12 + Log_line_number, Col_PrechargeExitDuration, int(duration_ph3), "Charging IOP", IOP_file)

    Max_current_ph4 = None
    for i in range(EndPh3, max_row + 1, 1):
        if abs(ws2[GetCol(Col_current) + str(i)].value) > precharge_end:
            EndPh4 = i
            break
        # Current calculation
        current_ph4 = 0 if abs(float(ws2[GetCol(Col_current) + str(i)].value)) == None else abs(
            float(ws2[GetCol(Col_current) + str(i)].value))
        current_ph4 *= 1000
        if Max_current_ph4 is None or current_ph4 > Max_current_ph4: Max_current_ph4 = current_ph4
    # EF.FillCell(12+Log_number, 12, Max_ph4, "Charging IOP", IOP_file)

    Max_current_ph5 = None
    Max_power_ph5 = None
    for i in range(EndPh4, max_row + 1, 1):
        if abs(ws2[GetCol(Col_current) + str(i)].value) < precharge_start:
            EndPh5 = i
            break
        # Current calculation
        current_ph5 = 0 if abs(float(ws2[GetCol(Col_current) + str(i)].value)) == None else abs(
            float(ws2[GetCol(Col_current) + str(i)].value))
        current_ph5 *= 1000
        if Max_current_ph5 is None or current_ph5 > Max_current_ph5: Max_current_ph5 = current_ph5
        # Power calculation
        power_ph5 = 0 if abs(float(ws2[GetCol(Col_power) + str(i)].value)) == None else abs(
            float(ws2[GetCol(Col_power) + str(i)].value))  # to prevent the program to tak a "None" value
        if Max_power_ph5 is None or power_ph5 > Max_power_ph5: Max_power_ph5 = power_ph5
    FillCell(12 + Log_line_number, Col_CurretAfterFlip, Max_current_ph5, "Charging IOP", IOP_file)
    FillCell(12 + Log_line_number, Col_PowerAfterFlip, Max_power_ph5, "Charging IOP",
                IOP_file)  # Fill the curret cell in Charging IOP sheet with Max_ph2