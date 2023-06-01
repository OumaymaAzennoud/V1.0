# Author :    Oumayma Azennoud
# Date:       31/05/2023
# Function:   Create phases
# ----------------------------------------------------------------------------------------------------------------------------------------------
# Import fonctions
# ----------------------------------------------------------------------------------------------------------------------------------------------
import Excel_fonctions as EF
from Excel_fonctions import max_row, max_col
# ----------------------------------------------------------------------------------------------------------------------------------------------
# General libreries
# ----------------------------------------------------------------------------------------------------------------------------------------------
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
import os

# ----------------------------------------------------------------------------------------------------------------------------------------------
# Global Variables
# ----------------------------------------------------------------------------------------------------------------------------------------------
#CSV2TABLE-------------------
Col_time=1,
Col_voltage=2
Col_current = 3
Col_power = 4
#Charging IOP----------------
Col_Nlog=3
Col_Type=4
Col_ChargingType=5
Col_USBVersion=6
Col_EDTRef=7
Col_TypeOfProduct=8
Col_Brand=9
Col_Model=10
Col_MaxCurrent=11
Col_Led=12
Col_PrechargeCrurrent=13
Col_PrechargeEndVoltage=14
Col_PrechargeDuration=15
Col_NormalChargeCurrent=16
Col_NormalChargePower=17
Col_PrechargeExitDuration=18
Col_CurretAfterFlip=19
Col_PowerAfterFlip=20
Col_VoltageNormalCharge=21
Col_TotalChargeDuration=22
Col_PassRemarkFail=23
Col_Sample=24
Col_Cable=25
Col_Comment=26
#-------------------------------
Log_number=10
IOP_file="Template"
Table_file="CSV2Table"
precharge_start=0.05
precharge_end=0.2
# ----------------------------------------------------------------------------------------------------------------------------------------------
# Workspaces and worksheets
# ----------------------------------------------------------------------------------------------------------------------------------------------
wb1 = load_workbook(
    'C:/Users/oumayma.azennoud/OneDrive - Eurofins Digital Testing International/Desktop/Python/Template.xlsx')
ws1 = wb1['Charging IOP']

wb2 = load_workbook(
    'C:/Users/oumayma.azennoud/OneDrive - Eurofins Digital Testing International/Desktop/Python/CSV2Table.xlsx')
ws2 = wb2['Sheet1']
# ----------------------------------------------------------------------------------------------------------------------------------------------

EF.Import_CSV(14)

Max_ph1 = None
for i in range(2, max_row + 1, 1):
    if abs(ws2[EF.GetCol(Col_current) + str(i)].value) > precharge_start:
        EndPh1 = i
        break
    current_ph1 = 0 if abs(float(ws2[EF.GetCol(Col_current) + str(i)].value)) == None else abs(float(ws2[EF.GetCol(Col_current) + str(i)].value))
    current_ph1 *= 1000
    if Max_ph1 is None or current_ph1 > Max_ph1: Max_ph1 = current_ph1
#EF.FillCell(12+Log_number,12, Max_ph1, "Charging IOP", Table_file)

Max_ph2 = None  #initialize the variable
for i in range(EndPh1, max_row + 1, 1): #From the start of the phase (end of last phase) to the last cell with data in the column, the loop will break before that as soon as the next phase is detected
    if abs(ws2[EF.GetCol(Col_current) + str(i)].value) > precharge_end:   #detect the next phase
        EndPh2 = i  #memorise the start row of the next phase
        break

    current_ph2 = 0 if abs(float(ws2[EF.GetCol(Col_current) + str(i)].value)) == None else abs(float(ws2[EF.GetCol(Col_current) + str(i)].value)) #to prevent the program to tak a "None" value
    current_ph2 *= 1000 #convert into mA
    if Max_ph2 is None or current_ph2 > Max_ph2: Max_ph2 = current_ph2  #select the max current value of the phase
EF.FillCell(12+Log_number,Col_PrechargeCrurrent, Max_ph2, "Charging IOP", IOP_file)     #Fill the curret cell in Charging IOP sheet with Max_ph2

Max_ph3 = None
for i in range(EndPh2, max_row + 1, 1):
    if abs(ws2[EF.GetCol(Col_current) + str(i)].value) < precharge_start:
        EndPh3 = i
        break
    current_ph3 = 0 if abs(float(ws2[EF.GetCol(Col_current) + str(i)].value)) == None else abs(float(ws2[EF.GetCol(Col_current) + str(i)].value))
    current_ph3 *= 1000
    if Max_ph3 is None or current_ph3 > Max_ph3: Max_ph3 = current_ph3
EF.FillCell(12+Log_number, Col_NormalChargeCurrent, Max_ph3, "Charging IOP", IOP_file)

Max_ph4 = None
for i in range(EndPh3, max_row + 1, 1):
    if abs(ws2[EF.GetCol(Col_current) + str(i)].value) > precharge_end:
        EndPh4 = i
        break
    current_ph4 = 0 if abs(float(ws2[EF.GetCol(Col_current) + str(i)].value)) == None else abs(float(ws2[EF.GetCol(Col_current) + str(i)].value))
    current_ph4 *= 1000
    if Max_ph4 is None or current_ph4 > Max_ph4: Max_ph4 = current_ph4
#EF.FillCell(12+Log_number, 12, Max_ph4, "Charging IOP", IOP_file)

Max_ph5 = None
for i in range(EndPh4, max_row + 1, 1):
    if abs(ws2[EF.GetCol(Col_current) + str(i)].value) < precharge_start:
        EndPh5 = i
        break
    current_ph5 = 0 if abs(float(ws2[EF.GetCol(Col_current) + str(i)].value)) == None else abs(float(ws2[EF.GetCol(Col_current) + str(i)].value))
    current_ph5 *= 1000
    if Max_ph5 is None or current_ph5 > Max_ph4: Max_ph5 = current_ph5
EF.FillCell(12+Log_number,Col_CurretAfterFlip, Max_ph5, "Charging IOP", IOP_file)